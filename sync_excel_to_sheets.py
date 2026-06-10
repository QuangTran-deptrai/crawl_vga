import os
import json
import httpx
import asyncio
import openpyxl
from dotenv import load_dotenv

load_dotenv()

async def upload_existing_data():
    webhook_url = os.getenv("GOOGLE_SHEETS_WEBHOOK_URL")
    if not webhook_url:
        print("Lỗi: Chưa có GOOGLE_SHEETS_WEBHOOK_URL trong file .env")
        return
        
    excel_path = "vga_data.xlsx"
    if not os.path.exists(excel_path):
        print(f"Lỗi: Không tìm thấy file {excel_path}")
        return
        
    print(f"Đang đọc file {excel_path}...")
    # Dùng data_only=True để lấy giá trị (value) thay vì công thức (formula)
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    if "GPU Lookup" not in wb.sheetnames:
        print("Không tìm thấy sheet 'GPU Lookup'")
        return
        
    sheet = wb["GPU Lookup"]
    
    rows = []
    # Lấy cả dòng tiêu đề
    for row in sheet.iter_rows(values_only=True):
        # Chuyển None thành chuỗi rỗng
        cleaned_row = [str(cell) if cell is not None else "" for cell in row]
        # Không đẩy lên các dòng trống hoàn toàn
        if any(cleaned_row):
            rows.append(cleaned_row)
            
    if not rows:
        print("Sheet trống, không có dữ liệu để đẩy lên.")
        return
        
    print(f"Tìm thấy {len(rows)} dòng dữ liệu. Đang đẩy lên Google Sheets...")
    
    payload = {
        "clear": True, # Xóa data cũ trên Sheets (nếu có) trước khi ghi
        "rows": rows
    }
    
    try:
        async with httpx.AsyncClient() as client:
            response = await client.post(webhook_url, json=payload, timeout=30.0)
            if response.status_code in [200, 302]: # Google Apps Script trả về 302 redirect
                print("Đẩy dữ liệu cũ lên Google Sheets thành công!")
            else:
                print(f"Lỗi khi đẩy lên Sheets: HTTP {response.status_code}")
                print(response.text)
    except Exception as e:
        print(f"Lỗi kết nối: {e}")

if __name__ == "__main__":
    asyncio.run(upload_existing_data())
