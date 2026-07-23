import os
import re
import json
import asyncio
import pandas as pd
from datetime import datetime, timezone, timedelta
from urllib.parse import urlparse, urljoin
from dotenv import load_dotenv
from playwright.async_api import async_playwright
import httpx
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
import google.generativeai as genai

load_dotenv()

class VGAScraper:
    def __init__(self):
        self.telegram_bot_token = os.getenv("TELEGRAM_BOT_TOKEN")
        self.telegram_chat_id = os.getenv("TELEGRAM_CHAT_ID")
        self.google_sheets_webhook_url = os.getenv("GOOGLE_SHEETS_WEBHOOK_URL")
        
        # API config: (domain, collection_handle, source_name)
        self.api_sources = [
            ("gearvn.com", "vga-rtx-50-series", "GearVN"),
            ("gearvn.com", "vga-card-man-hinh", "GearVN"),
            ("tinhocngoisao.com", "card-man-hinh", "Tin Học Ngôi Sao"),
        ]
        
        # Browser fallback URLs (chỉ dùng khi API thất bại)
        self.gearvn_urls = [
            "https://gearvn.com/collections/vga-rtx-50-series",
            "https://gearvn.com/collections/vga-card-man-hinh"
        ]
        self.thns_url = "https://tinhocngoisao.com/collections/card-man-hinh"
        
        self.data = []
        self.seen_urls = set()  # Track product URLs to prevent duplicates across crawl calls
    async def send_telegram_alert(self, message: str):
        if not self.telegram_bot_token or not self.telegram_chat_id:
            print(f"[Telegram Not Configured] {message}")
            return
            
        url = f"https://api.telegram.org/bot{self.telegram_bot_token}/sendMessage"
        payload = {
            "chat_id": self.telegram_chat_id,
            "text": message
        }
        try:
            async with httpx.AsyncClient() as client:
                response = await client.post(url, json=payload)
                if response.status_code != 200:
                    print(f"Failed to send telegram alert: {response.text}")
        except Exception as e:
            print(f"Error sending telegram alert: {e}")

    def clean_price(self, price_str):
        if not price_str:
            return 0
        cleaned = re.sub(r'[^\d]', '', price_str)
        return int(cleaned) if cleaned else 0

    def normalize_url(self, href, base_domain):
        """Normalize product URL: handle both absolute/relative hrefs, strip query/fragment."""
        if not href:
            return ""
        href = href.strip()
        if href.startswith('http'):
            full_url = href
        else:
            full_url = f"https://{base_domain}{href}" if href.startswith('/') else f"https://{base_domain}/{href}"
        # Strip query params & fragment, normalize trailing slash
        parsed = urlparse(full_url)
        clean = f"{parsed.scheme}://{parsed.netloc}{parsed.path.rstrip('/')}"
        return clean

    async def crawl_via_api(self, domain, collection_handle, source_name):
        """Crawl sản phẩm qua Shopify/Haravan JSON API. Nhanh hơn và đáng tin cậy hơn browser."""
        base_url = f"https://{domain}/collections/{collection_handle}/products.json"
        print(f"[API] Crawling {source_name}: {base_url}")
        
        page_num = 1
        total_added = 0
        
        try:
            async with httpx.AsyncClient(
                headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"},
                timeout=30.0,
                follow_redirects=True
            ) as client:
                while True:
                    url = f"{base_url}?limit=250&page={page_num}"
                    response = await client.get(url)
                    
                    if response.status_code != 200:
                        raise Exception(f"API trả HTTP {response.status_code} tại {url}")
                    
                    data = response.json()
                    products = data.get("products", [])
                    
                    if not products:
                        break
                    
                    for product in products:
                        handle = product.get("handle", "")
                        title = product.get("title", "").strip()
                        
                        if not title or not handle:
                            continue
                        
                        product_url = f"https://{domain}/products/{handle}"
                        
                        if product_url in self.seen_urls:
                            continue
                        
                        # Lấy giá từ variant đầu tiên
                        variants = product.get("variants", [])
                        if variants:
                            variant = variants[0]
                            discount_price = int(variant.get("price", 0) or 0)
                            compare_price = variant.get("compare_at_price")
                            if compare_price and int(compare_price) > 0:
                                original_price = int(compare_price)
                            else:
                                original_price = discount_price
                        else:
                            discount_price = 0
                            original_price = 0
                        
                        self.seen_urls.add(product_url)
                        self.data.append({
                            "source": source_name,
                            "raw_name": title,
                            "original_price": original_price,
                            "discount_price": discount_price,
                            "url": product_url
                        })
                        total_added += 1
                    
                    print(f"  [API] Trang {page_num}: {len(products)} sản phẩm, thêm mới: {total_added}")
                    page_num += 1
                    await asyncio.sleep(0.5)  # Rate limit nhẹ
            
            print(f"  [API] Tổng {source_name} ({collection_handle}): {total_added} sản phẩm")
            return True  # Thành công
            
        except Exception as e:
            print(f"  [API] Lỗi {source_name} ({collection_handle}): {e}")
            return False  # Thất bại, cần fallback

    def load_regex_rules(self, excel_path="vga_data.xlsx"):
        brand_rules = {}
        chipset_rules = {}
        vram_rules = {}
        if os.path.exists(excel_path):
            try:
                df = pd.read_excel(excel_path, sheet_name='GPU ref', engine='openpyxl')
                if 'GPU brand text' in df.columns and 'GPU brand' in df.columns:
                    for text, val in zip(df['GPU brand text'], df['GPU brand']):
                        if pd.notna(text) and pd.notna(val):
                            brand_rules[str(text).strip()] = str(val).strip()
                if 'Chipset text' in df.columns and 'Chipset' in df.columns:
                    for text, val in zip(df['Chipset text'], df['Chipset']):
                        if pd.notna(text) and pd.notna(val):
                            chipset_rules[str(text).strip()] = str(val).strip()
                if 'Video memory text' in df.columns and 'Video memory' in df.columns:
                    for text, val in zip(df['Video memory text'], df['Video memory']):
                        if pd.notna(text) and pd.notna(val):
                            vram_rules[str(text).strip()] = str(val).strip()
            except Exception as e:
                print(f"Error loading GPU ref: {e}")
        return brand_rules, chipset_rules, vram_rules

    def get_matches(self, text, rules_dict):
        text_upper = text.upper()
        
        # Tìm tất cả vị trí match trong text
        found = []  # [(start, end, key, value)]
        for key, val in rules_dict.items():
            key_upper = key.upper()
            start = 0
            while True:
                pos = text_upper.find(key_upper, start)
                if pos == -1:
                    break
                found.append((pos, pos + len(key_upper), key_upper, val))
                start = pos + 1
        
        # Sắp xếp: dài nhất trước, nếu bằng nhau thì vị trí sớm hơn trước
        found.sort(key=lambda x: (-(x[1] - x[0]), x[0]))
        
        # Loại bỏ các match bị overlap bởi match dài hơn
        kept = []
        for item in found:
            s, e, k, v = item
            is_overlapped = False
            for ks, ke, kk, kv in kept:
                # Nếu match hiện tại nằm gọn trong hoặc overlap với match đã giữ
                if s >= ks and e <= ke:
                    is_overlapped = True
                    break
                # Overlap một phần: match hiện tại bắt đầu trong match đã giữ
                if ks <= s < ke or ks < e <= ke:
                    is_overlapped = True
                    break
            if not is_overlapped:
                kept.append(item)
        
        # Trả về danh sách value duy nhất
        return list(set(v for _, _, _, v in kept))

    async def process_ambiguous_with_gemini(self, items_batch):
        api_key = os.getenv("GEMINI_API_KEY")
        if not api_key:
            print("No GEMINI_API_KEY found, skipping AI processing.")
            return []
        
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-2.5-flash', generation_config={"response_mime_type": "application/json"})
        
        prompt = "Extract brand, chipset, and vram_gb for the following GPU names. Return a JSON array of objects with keys: id, brand, chipset, vram_gb. If not found, return empty string.\n\n"
        prompt += json.dumps([{"id": item['id'], "raw_name": item['raw_name']} for item in items_batch], ensure_ascii=False)
        
        try:
            response = await model.generate_content_async(prompt)
            return json.loads(response.text)
        except Exception as e:
            print(f"Gemini API error: {e}")
            return []

    # ==================== BROWSER FALLBACK METHODS ====================
    # Các method dưới đây chỉ chạy khi JSON API thất bại

    async def _extract_gearvn_products(self, page, url, start_index=0):
        """Extract GearVN products from current page state, starting from start_index.
        
        Args:
            page: Playwright page object
            url: The collection URL being crawled
            start_index: Only extract products from this DOM index onward (for pagination dedup)
        """
        products = []
        blocks = page.locator('.proloop-block')
        count = await blocks.count()
        
        for i in range(start_index, count):
            block = blocks.nth(i)
            name_loc = block.locator('.proloop-name a')
            name = await name_loc.inner_text() if await name_loc.count() > 0 else ""
            
            product_url_path = await name_loc.get_attribute('href') if await name_loc.count() > 0 else ""
            product_url = self.normalize_url(product_url_path, "gearvn.com") if product_url_path else url
            
            # Skip if this product URL was already seen (cross-URL dedup)
            if product_url in self.seen_urls:
                continue
            
            highlight_loc = block.locator('.proloop-price--highlight')
            discount_price_str = await highlight_loc.inner_text() if await highlight_loc.count() > 0 else ""
            discount_price = self.clean_price(discount_price_str)
            
            compare_loc = block.locator('.proloop-price--compare del')
            if await compare_loc.count() > 0:
                original_price_str = await compare_loc.inner_text()
                original_price = self.clean_price(original_price_str)
            else:
                original_price = discount_price
                
            if name.strip():
                self.seen_urls.add(product_url)
                products.append({
                    "source": "GearVN",
                    "raw_name": name.strip(),
                    "original_price": original_price,
                    "discount_price": discount_price,
                    "url": product_url
                })
        return products

    async def crawl_gearvn_browser(self, page, url):
        print(f"Crawling GearVN: {url}")
        try:
            await page.goto(url, timeout=60000, wait_until="load")
            
            # Đợi sản phẩm thật load xong (AJAX dynamic loading)
            try:
                await page.wait_for_selector('.proloop-block .proloop-name', timeout=30000)
                print(f"Sản phẩm đã load xong tại {url}")
            except Exception:
                print(f"Timeout chờ sản phẩm load tại {url}, thử tiếp...")
                await page.wait_for_timeout(5000)
            
            # Dùng CSS ẩn luôn tất cả popup quảng cáo để không bao giờ bị che
            try:
                await page.add_style_tag(content=".modal, .modal-backdrop, #myModal, .collection-layout.js-loading::before { display: none !important; pointer-events: none !important; z-index: -1 !important; }")
                print("Đã chèn CSS ẩn popup quảng cáo GearVN.")
            except Exception:
                pass
            
            # Scrape trang hiện tại trước
            products = await self._extract_gearvn_products(page, url)
            self.data.extend(products)
            print(f"  Trang 1: {len(products)} sản phẩm (mới), bỏ qua trùng: seen_urls={len(self.seen_urls)}")
            
            # Xử lý pagination - Load More append thêm sản phẩm vào DOM
            selector_load_more = "#load_more"
            page_num = 1
            
            while True:
                try:
                    button = page.locator(selector_load_more)
                    if await button.count() > 0 and await button.is_visible():
                        # Đếm số sản phẩm hiện tại TRƯỚC khi click Load More
                        current_count = await page.locator('.proloop-block').count()
                        
                        await page.evaluate("document.querySelectorAll('.modal, .modal-backdrop, .js-loading').forEach(el => { el.style.pointerEvents = 'none'; el.classList.remove('js-loading'); })")
                        await page.evaluate("document.querySelector('#load_more')?.click()")
                        page_num += 1
                        
                        # Đợi sản phẩm mới load xong - đợi DOM có THÊM sản phẩm
                        try:
                            await page.wait_for_function(
                                f"""(prevCount) => {{
                                    const blocks = document.querySelectorAll('.proloop-block');
                                    if (blocks.length <= prevCount) return false;
                                    const lastBlock = blocks[blocks.length - 1];
                                    const name = lastBlock.querySelector('.proloop-name a');
                                    return name && name.innerText.trim().length > 0;
                                }}""",
                                arg=current_count,
                                timeout=15000
                            )
                        except Exception:
                            await page.wait_for_timeout(5000)
                        
                        # Chỉ extract sản phẩm MỚI từ index current_count trở đi
                        new_products = await self._extract_gearvn_products(page, url, start_index=current_count)
                        if not new_products:
                            print(f"  Trang {page_num}: không có sản phẩm mới, dừng.")
                            break
                        
                        self.data.extend(new_products)
                        print(f"  Trang {page_num}: {len(new_products)} sản phẩm mới")
                    else:
                        break
                except Exception as e:
                    print(f"  Lỗi pagination trang {page_num}: {e}")
                    break
            
            total_gearvn = sum(1 for d in self.data if d['source'] == 'GearVN')
            if total_gearvn == 0:
                await self.send_telegram_alert(f"Lỗi: GearVN không crawl được sản phẩm nào tại {url}")
            else:
                print(f"  Tổng GearVN từ {url}: {total_gearvn} sản phẩm")
                
        except Exception as e:
            await self.send_telegram_alert(f"Lỗi: GearVN lỗi kết nối hoặc xử lý tại {url}. Detail: {str(e)}")

    async def crawl_thns_browser(self, page, url):
        print(f"Crawling Tin Học Ngôi Sao: {url}")
        try:
            await page.goto(url, timeout=60000, wait_until="load")
            await page.wait_for_timeout(3000)
            
            # Dùng CSS ẩn popup/chat widget
            try:
                await page.add_style_tag(content=".modal, .modal-backdrop, [id*='onesignal'], [id*='fb-root'] { display: none !important; pointer-events: none !important; z-index: -1 !important; }")
            except Exception:
                pass
            
            # Xử lý pagination
            selector_load_more = ".btn-load__more"
            
            while True:
                try:
                    button = page.locator(selector_load_more)
                    if await button.count() > 0 and await button.is_visible():
                        current_count = await page.locator('.itemLoop').count()
                        await button.scroll_into_view_if_needed()
                        await page.evaluate("document.querySelectorAll('.modal, .modal-backdrop, [id*=\"onesignal\"]').forEach(el => el.remove())")
                        await button.click()
                        await page.wait_for_timeout(5000)
                        
                        new_count = await page.locator('.itemLoop').count()
                        if new_count == current_count:
                            break
                    else:
                        break
                except Exception as e:
                    break
                    
            blocks = page.locator('.itemLoop')
            count = await blocks.count()
            
            if count == 0:
                await self.send_telegram_alert(f"Lỗi: Tin Học Ngôi Sao không tìm thấy Selector .itemLoop hoặc lỗi kết nối tại {url}")
                return

            for i in range(count):
                block = blocks.nth(i)
                name_loc = block.locator('.pdLoopName a')
                name = await name_loc.inner_text() if await name_loc.count() > 0 else ""
                
                product_url_path = await name_loc.get_attribute('href') if await name_loc.count() > 0 else ""
                product_url = self.normalize_url(product_url_path, "tinhocngoisao.com") if product_url_path else url
                
                original_price = 0
                discount_price = 0
                
                # Giá bán (giá đã giảm) - nằm trong .item.price
                sale_price_loc = block.locator('.pdPrice .item.price')
                if await sale_price_loc.count() > 0:
                    sale_text = await sale_price_loc.inner_text()
                    discount_price = self.clean_price(sale_text)
                
                # Giá gốc (giá so sánh) - nằm trong .item.comparePrice
                compare_price_loc = block.locator('.pdPrice .item.comparePrice')
                if await compare_price_loc.count() > 0:
                    compare_text = await compare_price_loc.inner_text()
                    compare_val = self.clean_price(compare_text)
                    original_price = compare_val if compare_val > 0 else discount_price
                else:
                    original_price = discount_price

                if name and product_url not in self.seen_urls:
                    self.seen_urls.add(product_url)
                    self.data.append({
                        "source": "Tin Học Ngôi Sao",
                        "raw_name": name.strip(),
                        "original_price": original_price,
                        "discount_price": discount_price,
                        "url": product_url
                    })
        except Exception as e:
            await self.send_telegram_alert(f"Lỗi: Tin Học Ngôi Sao lỗi kết nối hoặc xử lý tại {url}. Detail: {str(e)}")
    async def async_run(self):
        # === Bước 1: Thử crawl qua JSON API (nhanh, không cần browser) ===
        api_failed_sources = []  # Track sources cần fallback browser
        
        for domain, handle, source_name in self.api_sources:
            success = await self.crawl_via_api(domain, handle, source_name)
            if not success:
                api_failed_sources.append((domain, handle, source_name))
        
        # === Bước 2: Fallback về browser cho sources bị lỗi API ===
        if api_failed_sources:
            print(f"\n⚠️ API thất bại cho {len(api_failed_sources)} nguồn, fallback về browser...")
            async with async_playwright() as p:
                browser = await p.chromium.launch(headless=True)
                context = await browser.new_context(
                    user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36"
                )
                page = await context.new_page()
                
                for domain, handle, source_name in api_failed_sources:
                    fallback_url = f"https://{domain}/collections/{handle}"
                    if source_name == "GearVN":
                        await self.crawl_gearvn_browser(page, fallback_url)
                    elif source_name == "Tin Học Ngôi Sao":
                        await self.crawl_thns_browser(page, fallback_url)
                
                await browser.close()
        else:
            print("\n✅ Tất cả nguồn đã crawl thành công qua API, không cần browser.")
            
        if not self.data:
            await self.send_telegram_alert("Cảnh báo: Không crawl được dữ liệu nào từ các trang.")
            return

        # Final dedup dựa trên URL sản phẩm
        seen = set()
        unique_data = []
        for item in self.data:
            key = item.get('url', '')
            if key and key not in seen:
                seen.add(key)
                unique_data.append(item)
            elif not key:
                unique_data.append(item)  # Giữ lại nếu không có URL
        
        dupes_removed = len(self.data) - len(unique_data)
        if dupes_removed > 0:
            print(f"⚠️ Đã loại bỏ {dupes_removed} sản phẩm trùng lặp (duplicate URL)")
        self.data = unique_data

        vn_tz = timezone(timedelta(hours=7))
        crawled_date_str = datetime.now(vn_tz).strftime("%d/%m/%Y %H:%M:%S")
        
        excel_path = "vga_data.xlsx"
        
        try:
            if os.path.exists(excel_path):
                wb = openpyxl.load_workbook(excel_path)
            else:
                wb = openpyxl.Workbook()
                sheet = wb.active
                sheet.title = "GPU Lookup"
                sheet.append(["source", "raw_name", "original_price", "discount_price", "url", "crawled_date", "brand", "chipset", "vram_gb"])
                
                ref_sheet = wb.create_sheet(title="GPU ref")
                ref_sheet.append(["GPU brand text", "GPU brand", "", "Chipset text", "Chipset", "", "Video memory text", "Video memory"])
                ref_sheet.append(["ASUS", "ASUS", "", "RTX 3060", "NVIDIA GeForce RTX 3060", "", "8GB", "8GB"])
                
                style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
                
                tab_brand = Table(displayName="GPU_brand", ref="A1:B2")
                tab_brand.tableStyleInfo = style
                ref_sheet.add_table(tab_brand)
                
                tab_chipset = Table(displayName="Chipset_Table", ref="D1:E2")
                tab_chipset.tableStyleInfo = style
                ref_sheet.add_table(tab_chipset)
                
                tab_vram = Table(displayName="VRAM_table", ref="G1:H2")
                tab_vram.tableStyleInfo = style
                ref_sheet.add_table(tab_vram)
                
            if "GPU Lookup" not in wb.sheetnames:
                wb.create_sheet("GPU Lookup")
            sheet = wb["GPU Lookup"]
            
            if sheet.max_row == 1 and sheet.cell(row=1, column=1).value is None:
                sheet.append(["source", "raw_name", "original_price", "discount_price", "url", "crawled_date", "brand", "chipset", "vram_gb"])
                
            brand_rules, chipset_rules, vram_rules = self.load_regex_rules(excel_path)
            
            ambiguous_items = []
            for idx, item in enumerate(self.data):
                item['id'] = idx
                item['is_ambiguous'] = False
                if brand_rules or chipset_rules or vram_rules:
                    b_matches = self.get_matches(item['raw_name'], brand_rules)
                    c_matches = self.get_matches(item['raw_name'], chipset_rules)
                    v_matches = self.get_matches(item['raw_name'], vram_rules)
                    if len(b_matches) > 1 or len(c_matches) > 1 or len(v_matches) > 1:
                        item['is_ambiguous'] = True
                        ambiguous_items.append(item)

            ai_results = {}
            if ambiguous_items:
                batch_size = 20
                for i in range(0, len(ambiguous_items), batch_size):
                    batch = ambiguous_items[i:i+batch_size]
                    res = await self.process_ambiguous_with_gemini(batch)
                    for r in res:
                        ai_results[r.get('id')] = r
                    if i + batch_size < len(ambiguous_items):
                        await asyncio.sleep(2)

            new_rows_for_sheets = []
            
            for item in self.data:
                row_idx = sheet.max_row + 1
                b_cell = f"B{row_idx}"
                
                if item['is_ambiguous'] and item['id'] in ai_results:
                    # Nhập nhằng (>1 kết quả) → dùng kết quả AI
                    ai_data = ai_results[item['id']]
                    brand_val = ai_data.get('brand', '')
                    chipset_val = ai_data.get('chipset', '')
                    vram_val = ai_data.get('vram_gb', '')
                    
                    sheet_brand_val = brand_val
                    sheet_chipset_val = chipset_val
                    sheet_vram_val = vram_val
                else:
                    # Regex trả đúng 1 kết quả → ghi text tĩnh
                    # Regex trả 0 kết quả → ghi công thức Excel để Excel tự tìm
                    b_matches = self.get_matches(item['raw_name'], brand_rules) if brand_rules else []
                    c_matches = self.get_matches(item['raw_name'], chipset_rules) if chipset_rules else []
                    v_matches = self.get_matches(item['raw_name'], vram_rules) if vram_rules else []
                    
                    brand_val = b_matches[0] if len(b_matches) == 1 else f'=_xlfn.TEXTJOIN(",",TRUE,_xlfn._xlws.FILTER(GPU_brand[GPU brand], (GPU_brand[GPU brand text]<>"") * ISNUMBER(SEARCH(GPU_brand[GPU brand text], {b_cell})), ""))'
                    chipset_val = c_matches[0] if len(c_matches) == 1 else f'=_xlfn.TEXTJOIN(",",TRUE,_xlfn._xlws.FILTER(Chipset_Table[Chipset], (Chipset_Table[Chipset text]<>"") * ISNUMBER(SEARCH(Chipset_Table[Chipset text], {b_cell})), ""))'
                    vram_val = v_matches[0] if len(v_matches) == 1 else f'=_xlfn.TEXTJOIN(",",TRUE,_xlfn._xlws.FILTER(VRAM_table[Video memory], (VRAM_table[Video memory text]<>"") * ISNUMBER(SEARCH(VRAM_table[Video memory text], {b_cell})), ""))'
                    
                    # Công thức riêng cho Google Sheets (dùng ; thay , cho locale VN)
                    sheet_brand_val = b_matches[0] if len(b_matches) == 1 else f'=IFERROR(TEXTJOIN(","; TRUE; FILTER(\'GPU ref\'!B:B; \'GPU ref\'!A:A<>""; ISNUMBER(SEARCH(\'GPU ref\'!A:A; {b_cell})))); "")'
                    sheet_chipset_val = c_matches[0] if len(c_matches) == 1 else f'=IFERROR(TEXTJOIN(","; TRUE; FILTER(\'GPU ref\'!E:E; \'GPU ref\'!D:D<>""; ISNUMBER(SEARCH(\'GPU ref\'!D:D; {b_cell})))); "")'
                    sheet_vram_val = v_matches[0] if len(v_matches) == 1 else f'=IFERROR(TEXTJOIN(","; TRUE; FILTER(\'GPU ref\'!H:H; \'GPU ref\'!G:G<>""; ISNUMBER(SEARCH(\'GPU ref\'!G:G; {b_cell})))); "")'
                
                row_data = [
                    item.get("source", ""),
                    item.get("raw_name", ""),
                    item.get("original_price", ""),
                    item.get("discount_price", ""),
                    item.get("url", ""),
                    crawled_date_str,
                    brand_val,
                    chipset_val,
                    vram_val
                ]
                sheet.append(row_data)
                
                sheet_row = [
                    item.get("source", ""),
                    item.get("raw_name", ""),
                    item.get("original_price", ""),
                    item.get("discount_price", ""),
                    item.get("url", ""),
                    crawled_date_str,
                    sheet_brand_val,
                    sheet_chipset_val,
                    sheet_vram_val
                ]
                new_rows_for_sheets.append(sheet_row)
                
            table_name = "GPU_Lookup_Data"
            table_ref = f"A1:I{sheet.max_row}"
            existing_table = None
            for tbl in sheet.tables.values():
                if tbl.displayName == table_name:
                    existing_table = tbl
                    break
            if existing_table:
                existing_table.ref = table_ref
            else:
                lookup_tab = Table(displayName=table_name, ref=table_ref)
                style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
                lookup_tab.tableStyleInfo = style
                sheet.add_table(lookup_tab)
                
            wb.save(excel_path)
            await self.send_telegram_alert(f"Thành công: Đã lấy được {len(self.data)} sản phẩm và lưu vào file Excel.")
            print(f"Data saved to {excel_path}")
            
            # Đẩy lên Google Sheets
            if self.google_sheets_webhook_url and new_rows_for_sheets:
                print(f"Đang đẩy {len(new_rows_for_sheets)} dòng lên Google Sheets...")
                payload = {
                    "clear": False, # Chỉ thêm vào cuối bảng
                    "rows": new_rows_for_sheets
                }
                try:
                    async with httpx.AsyncClient() as client:
                        response = await client.post(self.google_sheets_webhook_url, json=payload, timeout=30.0)
                        if response.status_code in [200, 302]:
                            print("Đã đồng bộ lên Google Sheets thành công!")
                        else:
                            print(f"Lỗi đồng bộ Sheets: HTTP {response.status_code}")
                except Exception as e:
                    print(f"Lỗi khi kết nối tới Webhook Google Sheets: {e}")
                    
            
        except PermissionError:
            await self.send_telegram_alert(f"Lỗi Permission Denied: Vui lòng đóng file {excel_path} trước khi chạy script.")
            print(f"Lỗi Permission Denied: Vui lòng đóng file {excel_path} trước khi chạy script.")
        except Exception as e:
            await self.send_telegram_alert(f"Lỗi khi lưu file Excel: {e}")
            print(f"Error saving to Excel: {e}")

    def run(self):
        asyncio.run(self.async_run())

if __name__ == "__main__":
    scraper = VGAScraper()
    scraper.run()
